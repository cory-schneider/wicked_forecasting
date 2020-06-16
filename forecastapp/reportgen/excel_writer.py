from forecastapp.reportgen import utils
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule
import os
from flask import (current_app)
from forecastapp.reportgen.utils import unique

def excel_writer(report_units, forecast_report):

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "ForecastWorksheet"

    date_list = forecast_report.date_list
    today = forecast_report.today

    header = report_units.pop(0)

    header_labels = []
    for i in header[0:5]:
        header_labels.append(i)
    header_labels.append("WEEKLY ROS")
    header_labels.append("DOH Target")
    header_labels.append("Inventory (Units)")
    header_labels.append("DOH")
    header_labels.append("Units to Target")
    header_labels.append("Actual Order (Units)")
    header_labels.append("Rolling DOH")
    header_labels.append("Units to Target")
    for i in date_list[2:]:
        header_labels.append("Actual Order (Units)")
        header_labels.append("WWB Sugg. Order (Units)")
        header_labels.append("Rolling DOH")
        header_labels.append("Units to Target")
    print(len(header_labels))

    header_dates = []
    for i in range(7):
        header_dates.append("")
    for i in date_list[:2]:
        header_dates.append(i)
        header_dates.append("")
        header_dates.append("")
    for i in date_list[2:]:
        header_dates.append(i)
        header_dates.append("")
        header_dates.append("")
        header_dates.append("")

    units_tgt_formula = "=($G3-I3)*$E3"
    units_tgt_ref = "J3"
    inv_doh_formula  = "=H3/E3"
    inv_doh_ref = "I3"
    first_doh_formula = "=IF((I3+(K3/$E3))-(K$1-H$1)>0,(I3+(K3/$E3))-(K$1-H$1),0)"
    first_doh_ref = "L3"
    second_doh_formula = "=IF((L3+(O3/$E3))-(N$1-K$1)>0,(L3+(O3/$E3))-(N$1-K$1),0)"
    second_doh_ref = "P3"
    doh_formula = "=IF((P3+(S3/$E3))-(R$1-N$1)>0,(P3+(S3/$E3))-(R$1-N$1),0)"
    doh_ref = "T3"

# Starting row in the spreadsheet for WSLR/ITEM combos.
    row_n = 3
    for line in report_units:
        row_ref = str(row_n)
        ws1['A'+row_ref] = line[0] # WSLR State
        ws1['B'+row_ref] = line[1] # WSLR Name
        ws1['C'+row_ref] = line[2] # WSLR ID
        pdcn = str(line[3])
        pdcn_verbose = utils.pdcn_plus_product_name(pdcn)
        ws1['D'+row_ref] = pdcn_verbose # PDCN
        ws1['E'+row_ref] = line[4] # Daily ROS
        ws1['F'+row_ref] = "=E"+row_ref+"*7" # Weekly ROS
        ws1['G'+row_ref] = 21 # DOH Target
        ws1['H'+row_ref] = line[5] # Inventory (Units)
        ws1['H'+row_ref].border = Border(left=Side(style='thin'))
        ws1['I'+row_ref] = Translator(inv_doh_formula, inv_doh_ref).translate_formula("I"+row_ref) # Inventory DOH
        ws1['J'+row_ref] = Translator(units_tgt_formula, units_tgt_ref).translate_formula("J"+row_ref) # Inventory Units to Tgt
        ws1['K'+row_ref] = line[6] # First Order
        ws1['K'+row_ref].border = Border(left=Side(style='thin'))
        ws1['L'+row_ref] = Translator(first_doh_formula, first_doh_ref).translate_formula("L"+row_ref) # First Order DOH
        ws1['M'+row_ref] = Translator(units_tgt_formula, units_tgt_ref).translate_formula("M"+row_ref) # First Order DOH
        ws1['N'+row_ref] = line[7] # Second Order
        ws1['N'+row_ref].border = Border(left=Side(style='thin'))
        ws1['O'+row_ref] = line[7] # Second WWB Sugg. Order
        ws1['P'+row_ref] = Translator(second_doh_formula, second_doh_ref).translate_formula("P"+row_ref) # Second Order DOH
        ws1['Q'+row_ref] = Translator(units_tgt_formula, units_tgt_ref).translate_formula("Q"+row_ref) # Second Order Units to Target
# Column R is 18.
        col_n = 18
        for order in line[8:]:
            ws1[get_column_letter(col_n)+row_ref] = order # Next
            ws1[get_column_letter(col_n)+row_ref].border = Border(left=Side(style='thin'))
            ws1[get_column_letter(col_n+1)+row_ref] = order # Next
            ws1[get_column_letter(col_n+2)+row_ref] = Translator(doh_formula, doh_ref).translate_formula(get_column_letter(col_n+2)+row_ref) # Next Order DOH
            ws1[get_column_letter(col_n+3)+row_ref] = Translator(units_tgt_formula, units_tgt_ref).translate_formula(get_column_letter(col_n+3)+row_ref) # Next Order Units to Target
            col_n += 4

        row_n += 1

    last_cell = f"{get_column_letter(col_n)}{row_n}"

# Styling

    ws1.auto_filter.ref = "A2:AS2"
    ws1.freeze_panes = 'F3'

    conditional_cols = "I L P T X AB AF AJ AN AR".split()

    redFill = PatternFill(start_color='fcc7c9',
                          end_color='fcc7c9',
                          fill_type='solid')
    for cond_col in conditional_cols:
        ws1.conditional_formatting.add(f'{cond_col}3:{cond_col}{row_n}',
                    CellIsRule(operator='lessThan', formula=['7'], fill=redFill))
        ws1.conditional_formatting.add(f'{cond_col}3:{cond_col}{row_n}',
                    CellIsRule(operator='greaterThan', formula=['35'], fill=redFill))

    for i, rowOfCellObjects in enumerate(ws1['H3':last_cell]):
        for n, cellObj in enumerate(rowOfCellObjects):
            cellObj.number_format = '#0'

    column_widths = [7.6, 35, 18, 45, 15, 8.5, 8,
                     11, 7.5, 8,
                     13, 8, 9,
                     13, 14.5, 8, 9, 13, 14.5, 8, 9,
                     13, 14.5, 8, 9, 13, 14.5, 8, 9,
                     13, 14.5, 8, 9, 13, 14.5, 8, 9,
                     13, 14.5, 8, 9, 13, 14.5, 8, 9,]
    col_count = 1
    for width in column_widths:
        ws1.column_dimensions[get_column_letter(col_count)].width = width
        col_count += 1

    red_zone_style = PatternFill(start_color="ff9896", end_color="ff9896", fill_type = "solid")
    for i, rowOfCellObjects in enumerate(ws1['H1':'M2']):
        for n, cellObj in enumerate(rowOfCellObjects):
            cellObj.fill = red_zone_style

    yellow_zone_style = PatternFill(start_color="ffdd71", end_color="ffdd71", fill_type = "solid")
    for i, rowOfCellObjects in enumerate(ws1['N1':'Y2']):
        for n, cellObj in enumerate(rowOfCellObjects):
            cellObj.fill = yellow_zone_style

    green_zone_style = PatternFill(start_color="98df8a", end_color="98df8a", fill_type = "solid")
    for i, rowOfCellObjects in enumerate(ws1['Z1':'AS2']):
        for n, cellObj in enumerate(rowOfCellObjects):
            cellObj.fill = green_zone_style

    thick_bottom_border = Border(left=Side(style=None),
                          right=Side(style=None),
                          top=Side(style=None),
                          bottom=Side(style='thick'))
    for i, rowOfCellObjects in enumerate(ws1['A2':'AS2']):
        for n, cellObj in enumerate(rowOfCellObjects):
            cellObj.border = thick_bottom_border

    header_font = Font(bold=True)
    for i, rowOfCellObjects in enumerate(ws1['A1':'AS2']):
        for n, cellObj in enumerate(rowOfCellObjects):
            cellObj.font = header_font

    col_n = 1
    for item in header_dates:
        ws1.cell(row=1, column=col_n).value = item
        ws1.cell(row=1, column=col_n).alignment = Alignment(wrap_text=True)
        col_n += 1
    col_n = 1
    for item in header_labels:
        ws1.cell(row=2, column=col_n).value = item
        ws1.cell(row=2, column=col_n).alignment = Alignment(wrap_text=True)
        col_n += 1

    dest_filename = forecast_report.temp_report_path + '/ForecastWorksheet.xlsx'
    wb.save(filename = dest_filename)

    return dest_filename
